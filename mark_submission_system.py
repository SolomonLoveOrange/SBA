import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import time
import mysql.connector
from mysql.connector import Error
import openpyxl
from openpyxl import Workbook
import os
import matplotlib.pyplot as plt
from collections import defaultdict

# Function to get or insert language into the database
def get_or_insert_language(cursor, language):
    # Check if the language already exists (case-insensitive)
    query = "SELECT lang_id FROM language WHERE LOWER(language) = LOWER(%s)"
    cursor.execute(query, (language,))
    result = cursor.fetchone()
    
    if result:
        # If the language exists, return its ID
        return result[0]
    else:
        # If the language does not exist, insert it and return the new ID
        insert_query = "INSERT INTO language (language) VALUES (%s)"
        cursor.execute(insert_query, (language,))
        return cursor.lastrowid

# Function to get or insert elective into the database
def get_or_insert_elective(cursor, elective):
    # Check if the elective already exists (case-insensitive)
    query = "SELECT elective_id FROM student_elective WHERE LOWER(elective) = LOWER(%s)"
    cursor.execute(query, (elective,))
    result = cursor.fetchone()
    
    if result:
        # If the elective exists, return its ID
        return result[0]
    else:
        # If the elective does not exist, insert it and return the new ID
        insert_query = "INSERT INTO student_elective (elective) VALUES (%s)"
        cursor.execute(insert_query, (elective,))
        return cursor.lastrowid

# Function to get or insert gender into the database
def get_or_insert_gender(cursor, gender):
    # Check if the gender already exists (case-insensitive)
    query = "SELECT gender_id FROM gender WHERE LOWER(gender) = LOWER(%s)"
    cursor.execute(query, (gender,))
    result = cursor.fetchone()
    
    if result:
        # If the gender exists, return its ID
        return result[0]
    else:
        # If the gender does not exist, insert it and return the new ID
        insert_query = "INSERT INTO gender (gender) VALUES (%s)"
        cursor.execute(insert_query, (gender,))
        return cursor.lastrowid
    
def drop_and_create_database(host, user, password, database_name):
    try:
        # Connect to MySQL server
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password
        )

        cursor = conn.cursor(buffered=True)  # Buffered cursor for larger datasets

        # Drop the database if it exists
        drop_db_query = f"DROP DATABASE IF EXISTS {database_name}"
        cursor.execute(drop_db_query)
        print(f"Dropped database '{database_name}' if it existed.")

        # Create the database
        create_db_query = f"CREATE DATABASE {database_name}"
        cursor.execute(create_db_query)
        print(f"Created database '{database_name}'.")

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        # Close cursor and connection
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()

def create_fresh_db_tables(host, user, password, database_name):
    try:
        # Connect to MySQL server
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database=database_name
        )

        cursor = conn.cursor()

        # SQL script to set up database schema
        query = """
        
        
        DROP TABLE IF EXISTS gender;

        
        CREATE TABLE IF NOT EXISTS `gender` (
          `gender_id` int(11) NOT NULL AUTO_INCREMENT,
          `gender` char(2) NOT NULL,
          PRIMARY KEY (`gender_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        
        DROP TABLE IF EXISTS language;
        CREATE TABLE IF NOT EXISTS `language` (
          `lang_id` int(11) NOT NULL AUTO_INCREMENT,
          `language` varchar(20) NOT NULL,
          PRIMARY KEY (`lang_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        
        DROP TABLE IF EXISTS student_elective;
        CREATE TABLE IF NOT EXISTS `student_elective` (
          `elective_id` int(11) NOT NULL AUTO_INCREMENT,
          `elective` varchar(5) NOT NULL,
          PRIMARY KEY (`elective_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        
        DROP TABLE IF EXISTS schools;
        CREATE TABLE IF NOT EXISTS `schools` (
          `school_id` int(11) NOT NULL AUTO_INCREMENT,
          `school` varchar(20) NOT NULL DEFAULT 'sch',
          PRIMARY KEY (`school_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;

        DROP TABLE IF EXISTS school_class;
        CREATE TABLE IF NOT EXISTS `school_class` (
          `schlass_id` int(11) NOT NULL AUTO_INCREMENT,
          `schlass_code` char(7) NOT NULL,
          `school_fk_id` int(11) NOT NULL,
          `schlass_lang_fk_id` int(11) NOT NULL,
          PRIMARY KEY (`schlass_id`),
          KEY `school_fk_id` (`school_fk_id`),
          KEY `schlass_lang_fk_id` (`schlass_lang_fk_id`),
          CONSTRAINT `school_class_ibfk_1` FOREIGN KEY (`school_fk_id`) REFERENCES `schools` (`school_id`),
          CONSTRAINT `school_class_ibfk_2` FOREIGN KEY (`schlass_lang_fk_id`) REFERENCES `language` (`lang_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;

        DROP TABLE IF EXISTS stud_school_class;
        CREATE TABLE IF NOT EXISTS `stud_school_class` (
          `student_id` int(11) NOT NULL AUTO_INCREMENT,
          `student_name` varchar(50) NOT NULL,
          `gender_fk_id` int(11) NOT NULL,
          `school_class_fk_id` int(11) NOT NULL,
          `stud_class_elective_fk_id` int(11) NOT NULL,
          PRIMARY KEY (`student_id`),
          KEY `gender_fk_id` (`gender_fk_id`),
          KEY `school_class_fk_id` (`school_class_fk_id`),
          KEY `stud_class_elective_fk_id` (`stud_class_elective_fk_id`),
          CONSTRAINT `stud_school_class_ibfk_1` FOREIGN KEY (`school_class_fk_id`) REFERENCES `school_class` (`schlass_id`),
          CONSTRAINT `stud_school_class_ibfk_2` FOREIGN KEY (`gender_fk_id`) REFERENCES `gender` (`gender_id`),
          CONSTRAINT `stud_school_class_ibfk_3` FOREIGN KEY (`stud_class_elective_fk_id`) REFERENCES `student_elective` (`elective_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;

        DROP TABLE IF EXISTS student_examp_score_paper_2;
        CREATE TABLE IF NOT EXISTS `student_exam_score_paper_2` (
          `paper_2_id` int(11) NOT NULL AUTO_INCREMENT,
          `Eq1` int NOT NULL,
          `Eq2` int NOT NULL,
          `Eq3` int NOT NULL,
          `Eq4` int NOT NULL,
          `student_id_fk` int(11) NOT NULL,
          PRIMARY KEY (`paper_2_id`),
          KEY `student_id_fk` (`student_id_fk`),
          CONSTRAINT `student_exam_score_paper_2_ibfk_1` FOREIGN KEY (`student_id_fk`) REFERENCES `stud_school_class` (`student_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;

        DROP TABLE IF EXISTS student_exam_score_section_a;
        CREATE TABLE IF NOT EXISTS `student_exam_score_section_a` (
          `id` int(11) NOT NULL AUTO_INCREMENT,
          `section_A_Mc` int NOT NULL,
          `student_id_fk` int(11) NOT NULL,
          PRIMARY KEY (`id`),
          KEY `student_id_fk` (`student_id_fk`),
          CONSTRAINT `student_exam_score_section_a_ibfk_1` FOREIGN KEY (`student_id_fk`) REFERENCES `stud_school_class` (`student_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;

        DROP TABLE IF EXISTS student_exam_score_section_b;
        CREATE TABLE IF NOT EXISTS `student_exam_score_section_b` (
          `section_B_id` int(11) NOT NULL AUTO_INCREMENT,
          `Bq1` int NOT NULL,
          `Bq2` int NOT NULL,
          `Bq3` int NOT NULL,
          `Bq4` int NOT NULL,
          `Bq5` int NOT NULL,
          `student_id_fk` int(11) NOT NULL,
          PRIMARY KEY (`section_B_id`),
          KEY `student_id_fk` (`student_id_fk`),
          CONSTRAINT `student_exam_score_section_b_ibfk_1` FOREIGN KEY (`student_id_fk`) REFERENCES `stud_school_class` (`student_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        
        DROP TABLE IF EXISTS `student_computed_grade`;
        CREATE TABLE `student_computed_grade` (
          `grade_id` int(11) NOT NULL AUTO_INCREMENT,
          `paper_1_grade` decimal(10,2) DEFAULT NULL,
          `paper_2_grade` decimal(10,2) DEFAULT NULL,
          `level_score` varchar(10) DEFAULT NULL,
          `grade` decimal(10,2) DEFAULT NULL,
          `stud_school_class_id` int(11) NOT NULL,
          PRIMARY KEY (`grade_id`),
          UNIQUE KEY `stud_school_class_id_2` (`stud_school_class_id`),
          KEY `stud_school_class_id` (`stud_school_class_id`),
          CONSTRAINT `student_computed_grade_ibfk_1` FOREIGN KEY (`stud_school_class_id`) REFERENCES `stud_school_class` (`student_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        """

        # Execute the entire SQL script
        cursor.execute(query,multi=True)
        print("Database setup successfully.")

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        # Close cursor and connection
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        
def test_connection(button):
    #test_connection_button.config(state="disabled")
    host = entry_host.get() 
    user = entry_user.get()
    password = entry_password.get()

    try:
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password
        )

        if conn.is_connected():
            entry_host.config(state='readonly')
            entry_user.config(state='readonly')
            entry_password.config(state='readonly')
            button.config(state=tk.DISABLED, text="Connection Established")
            conn.close()
        else:
            button.config(state=tk.NORMAL, text="Click To Retry...")
            messagebox.showerror("Connection Status", "Connection Failed")

    except Error as e:
        messagebox.showerror("Connection Status", f"Error: {str(e)}")

# Function to connect to MySQL and execute query
def execute_query(host, user, password, database_name, query):
    try:
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database=database_name
        )

        if conn.is_connected():
            cursor = conn.cursor(dictionary=True)
            cursor.execute(query)
            results = cursor.fetchall()
            cursor.close()
            conn.close()
            return results

    except Error as e:
        print(f"Error: {e}")

    return None


def plot_results(results,data_point='gender', output_folder = 'result'):
    if results:
        # Prepare data for plotting
        data_points = [result[f'{data_point}'] for result in results]
        counts = [result['count'] for result in results]

        # Plotting the pie chart
        plt.figure(figsize=(8, 6))
        plt.pie(counts, labels=[f"{data_point} ({count})" for data_point, count in zip(data_points, counts)], autopct='%1.1f%%', startangle=140)
        plt.title(f'Distribution of Students by {data_point} ')
        plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle

        # Save the pie chart to the output directory
        
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
    
        output_file = os.path.join(output_folder, f'{data_point}_distribution_piechart.png')
        plt.savefig(output_file)

        # Display the plot (optional)
        #plt.show()

        print(f"Pie chart saved to: {output_file}")
    else:
        print("No results returned from the query.")

def fetch_data_from_db(query, db_config):
    try:
        # Establish a database connection
        conn = mysql.connector.connect(
            host=db_config['host'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['database']
        )
        cursor = conn.cursor()

        # Execute the query
        cursor.execute(query)

        # Fetch all rows from the executed query
        rows = cursor.fetchall()

        # Fetch column names
        column_names = [i[0] for i in cursor.description]

        return column_names, rows

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return None, None

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

# Function to save rows to Excel files based on the first column value
def save_to_excel(column_names, rows, output_folder):
    # Group rows by the value of the first column
    grouped_rows = defaultdict(list)
    for row in rows:
        file_key = row[0]
        grouped_rows[file_key].append(row)

    # Save each group to a separate Excel file
    for file_key, group in grouped_rows.items():
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Append the column names as the first row
        ws.append(column_names[1:])

        # Append the rows in the group
        for row in group:
            ws.append(row[1:])

        # Save the workbook to the specified file
        output_file = os.path.join(output_folder, f"{file_key}.xlsx")
        wb.save(output_file)
        print(f"Data saved to {output_file}")
   

def select_input_folder():
    input_folder = filedialog.askdirectory()
    if input_folder:
        input_folder_var.set(input_folder)

def select_output_folder():
    output_folder = filedialog.askdirectory()
    if output_folder:
        output_folder_var.set(output_folder)


def perform_analysis(button):
    # Disable the button
    button.config(state=tk.DISABLED)
    
    input_folder = input_folder_var.get()
    output_folder = output_folder_var.get()
    if input_folder and output_folder:
        #messagebox.showinfo("Selected Folders", f"Input Folder: {input_folder}\nOutput Folder: {output_folder}")
        host= entry_host.get()
        user= entry_user.get()
        password= entry_password.get()
        database_name= 'mark_submission_system'
    
        drop_and_create_database(host, user, password, database_name)
        time.sleep(10) # sleep 10 seconds
        create_fresh_db_tables(host, user, password, database_name)
        time.sleep(10) # sleep 10 seconds
    

    
    
        # Initialize SQL Database Connection For Inserting Records
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password,  # Replace with the actual password
            database=database_name
        )

        try:
            cursor = conn.cursor(buffered=True)  # Buffered cursor for larger datasets

            folder_path = input_folder
            print(os.listdir(folder_path))
        
            for file_name in os.listdir(folder_path):
            
                if file_name.endswith(".xlsx"):
                    print(f"Inserting contents of file {file_name} into Mysql Database ...")
                    schoolname = file_name[:-5]
                    last_inserted_school_row_id = 0

                    # Insert school
                    sql_school = "INSERT INTO schools (school_id, school) VALUES (default, %s)"
                    cursor.execute(sql_school, (schoolname,))
                    last_inserted_school_row_id = cursor.lastrowid

                    # Load Excel workbook
                    wb = openpyxl.load_workbook(os.path.join(folder_path, file_name))
                    sheet = wb.active

                    # Define column mapping (adjust as needed)
                    column_mapping = {
                        1: "Classnum",
                        2: "Name",
                        3: "Gender",
                        4: "Elective",
                        5: "Lang",
                        6: "Mc", # total is 40 pts
                        
                        7: "Bq1", # n/12 * 5: on each question
                        8: "Bq2", # n/12 * 5: on each question
                        9: "Bq3", # n/12 * 5: on each question
                        10: "Bq4",# n/12 * 5: on each question
                        11: "Bq5",# n/12 * 5: on each question 
                        # Weighting of Paper I is 68.75% == (40 + 25) points
                        
                        12: "Eq1",# n/15 * 4 : on each question
                        13: "Eq2",# n/15 * 4 : on each question
                        14: "Eq3",# n/15 * 4 : on each question
                        15: "Eq4" # n/15 * 4 : on each question
                        # Weighting of Paper II is 31.25% === 16points
                    }

                    # Iterate through rows in the sheet
                    for i in range(2, sheet.max_row + 1):
                        row_data = {}
                        for j in range(1, sheet.max_column + 1):
                            cell_value = sheet.cell(row=i, column=j).value
                            if j in column_mapping:
                                row_data[column_mapping[j]] = cell_value if cell_value is not None and cell_value != "" else None
                            #print(cell_value if cell_value is not None else "None", end=" ")

                        # Insert language and retrieve foreign key
                        shclass_lang_fk_id = get_or_insert_language(cursor, row_data['Lang'])

                        # Insert school class
                        sql_school_class = """
                            INSERT INTO school_class (schlass_id, schlass_code, school_fk_id, schlass_lang_fk_id)
                            VALUES (default, %s, %s, %s)
                        """
                        cursor.execute(sql_school_class, (row_data["Classnum"], last_inserted_school_row_id, shclass_lang_fk_id))
                        last_inserted_schlass_row_id = cursor.lastrowid

                        # Insert student elective and retrieve foreign key
                        stud_class_elective_fk_id = get_or_insert_elective(cursor, row_data['Elective'])

                        # Insert gender and retrieve foreign key
                        gender_fk_id = get_or_insert_gender(cursor, row_data['Gender'])

                        # Insert student school class
                        sql_stud_school_class = """
                            INSERT INTO stud_school_class (student_id, student_name, gender_fk_id, school_class_fk_id, stud_class_elective_fk_id)
                            VALUES (default, %s, %s, %s, %s)
                        """
                        cursor.execute(sql_stud_school_class, (row_data["Name"], gender_fk_id, last_inserted_schlass_row_id, stud_class_elective_fk_id))
                        last_inserted_stud_school_class_id = cursor.lastrowid

                        # Insert student exam scores section A (MC)
                        sql_exam_section_a = """
                            INSERT INTO student_exam_score_section_a (id, section_A_Mc, student_id_fk)
                            VALUES (default, %s, %s)
                        """
                        cursor.execute(sql_exam_section_a, (row_data["Mc"], last_inserted_stud_school_class_id))

                        # Insert student exam scores section B (Bq1-Bq5)
                        sql_exam_section_b = """
                            INSERT INTO student_exam_score_section_b (section_B_id, Bq1, Bq2, Bq3, Bq4, Bq5, student_id_fk)
                            VALUES (default, %s, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(sql_exam_section_b, (row_data["Bq1"], row_data["Bq2"], row_data["Bq3"], row_data["Bq4"], row_data["Bq5"], last_inserted_stud_school_class_id))

                        PAPER_I_Percent_Score = row_data["Mc"] +  (row_data["Bq1"]/12 * 5) + (row_data["Bq2"]/12 * 5) + (row_data["Bq3"]/12 * 5) + (row_data["Bq4"]/12 * 5) + (row_data["Bq5"]/12 * 5) #n/12 * 5: on each question
                        # Weighting of Paper I is 68.75% == (40 + 25) points
                        PAPER_I_Percent_Score = PAPER_I_Percent_Score/65 * 68.75

                        




                        # Insert student exam scores paper 2 (Eq1-Eq4)
                        sql_exam_paper_2 = """
                            INSERT INTO student_exam_score_paper_2 (paper_2_id, Eq1, Eq2, Eq3, Eq4, student_id_fk)
                            VALUES (default, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(sql_exam_paper_2, (row_data["Eq1"], row_data["Eq2"], row_data["Eq3"], row_data["Eq4"], last_inserted_stud_school_class_id))

                        PAPER_II_Percent_Score = (row_data["Eq1"]/15 * 4) + (row_data["Eq2"]/15 * 4) + (row_data["Eq3"]/15 * 4) + (row_data["Eq4"]/15 * 4)  #n/15 * 4: on each question
                        # Weighting of Paper II is 31.25% == 16 points
                        PAPER_II_Percent_Score = PAPER_II_Percent_Score/16 * 31.25
                        

                        ### Insert Results of Paper I and Paper II to Database
                        LEVEL_SCORE = 'U'
                        GRADE = PAPER_I_Percent_Score + PAPER_II_Percent_Score
                        
                        if (GRADE >= 85):
                            LEVEL_SCORE = '5**'
                        elif(GRADE >= 77):
                            LEVEL_SCORE = '5*'
                        elif(GRADE >= 70):
                            LEVEL_SCORE = '5'
                        elif(GRADE >= 55):
                            LEVEL_SCORE = '4'
                        elif(GRADE >= 45):
                            LEVEL_SCORE = '3'
                        elif(GRADE >= 30):
                            LEVEL_SCORE = '2'
                        elif(GRADE >= 20):
                            LEVEL_SCORE = '1'
                        
                        

                        sql_exam_percentages = """
                            INSERT INTO student_computed_grade (grade_id, paper_1_grade, paper_2_grade,level_score,grade, stud_school_class_id)
                            VALUES (default, %s, %s,%s,%s, %s)
                        """
                        cursor.execute(sql_exam_percentages, (PAPER_I_Percent_Score,PAPER_II_Percent_Score,LEVEL_SCORE,GRADE, last_inserted_stud_school_class_id))




                        conn.commit()  # Commit per school

                        #print()  # Newline for clarity
                else:
                    messagebox.showerror("Doing Nothing", f"The folder {input_folder} contains no excel files. Ensure you provide the corrent upload folder  ")


            cursor.close()  # Close cursor after all operations
            ## Start of Generation of Reports
            # SQL Query to count male and female students
            query = """
                SELECT gender.gender, COUNT(*) AS count
                FROM stud_school_class
                INNER JOIN gender ON stud_school_class.gender_fk_id = gender.gender_id
                WHERE gender.gender IN ('M', 'F')
                GROUP BY gender.gender;
            """

            query_languages = """
                SELECT language.language, COUNT(*) AS count
                FROM stud_school_class
                INNER JOIN school_class ON stud_school_class.school_class_fk_id = school_class.schlass_id
                INNER JOIN language on school_class.schlass_lang_fk_id = language.lang_id
                WHERE language.language IN ('E', 'C')
                GROUP BY language.language;
            """

            query_electives = """
                SELECT student_elective.elective, COUNT(*) AS count
                FROM stud_school_class
                INNER JOIN student_elective ON stud_school_class.stud_class_elective_fk_id = student_elective.elective_id
                WHERE student_elective.elective IN ('AC', 'BC','AB')
                GROUP BY student_elective.elective;
            """

            # Execute the query, plot and save png files to output folder
            results = execute_query(host, user, password, database_name, query_electives)
            plot_results(results,'elective',output_folder=output_folder)
            results = execute_query(host, user, password, database_name, query)
            plot_results(results,'gender',output_folder=output_folder)
            results = execute_query(host, user, password, database_name, query_languages)
            plot_results(results,'language',output_folder=output_folder)


            print("Data Operations Completed Successfully.......")
        except mysql.connector.Error as err:
            print(f"Error: {err}")
            conn.rollback()  # Rollback in case of any error
        else:
            # create the results.xlsx and develop a plot
            # Database configuration
            db_config = {
                'host': host,
                'user': user,
                'password': password,
                'database': 'mark_submission_system'
            }
            # SQL query to fetch data
            query = """
                    SELECT schools.school, school_class.schlass_code, stud_school_class.student_name,
                       gender.gender, student_elective.elective, language.language, 
                       student_computed_grade.paper_1_grade, student_computed_grade.paper_2_grade, 
                       student_computed_grade.grade, student_computed_grade.level_score 
                    FROM schools 
		            INNER JOIN school_class on schools.school_id = school_class.school_fk_id
                    INNER JOIN  language on school_class.schlass_lang_fk_id = language.lang_id
                    INNER JOIN  stud_school_class on school_class.schlass_id = stud_school_class.school_class_fk_id
                    INNER JOIN  student_elective on stud_school_class.stud_class_elective_fk_id = student_elective.elective_id
                    INNER JOIN  gender on stud_school_class.gender_fk_id = gender.gender_id
                    INNER JOIN student_computed_grade on stud_school_class.student_id = student_computed_grade.stud_school_class_id
                    ORDER BY schools.school
                """


            # Fetch data from database
            column_names, rows = fetch_data_from_db(query, db_config)

            if column_names and rows:
                # Save data to Excel
                save_to_excel(column_names, rows,output_folder)
            else:
                print("Failed to fetch data from database.")

            
            

        finally:
            conn.close()  # Close connection at the end
            messagebox.showinfo("Task Complete", f"The Statistic Analysis task is complete.\n Access Report of Results in {output_folder} ")





    else:
        messagebox.showwarning("Incomplete Selection", "Please select both input and output folders.")
    button.config(state=tk.NORMAL)



def start_long_running_task():
    thread = threading.Thread(target=perform_analysis, args=(task_button,))
    thread.start()

def show_selected_folders():
    input_folder = input_folder_var.get()
    output_folder = output_folder_var.get()
    if input_folder and output_folder:
        messagebox.showinfo("Selected Folders", f"Input Folder: {input_folder}\nOutput Folder: {output_folder}")
    else:
        messagebox.showwarning("Incomplete Selection", "Please select both input and output folders.")

# Create the main window
root = tk.Tk()
root.title("Database and Folder Selection")

# Database connection widgets
tk.Label(root, text="Host:").grid(row=0, column=0, padx=10, pady=10)
entry_host = tk.Entry(root)
entry_host.insert(0, "localhost")  # Set default value to localhost
entry_host.grid(row=0, column=1, padx=10, pady=10)

tk.Label(root, text="Username:").grid(row=1, column=0, padx=10, pady=10)
entry_user = tk.Entry(root)
entry_user.insert(0, "root")  # Set default value to root
entry_user.grid(row=1, column=1, padx=10, pady=10)

tk.Label(root, text="Password:").grid(row=2, column=0, padx=10, pady=10)
entry_password = tk.Entry(root, show="*")
entry_password.grid(row=2, column=1, padx=10, pady=10)

test_connection_button = tk.Button(root, text="Test Connection", command=lambda: test_connection(test_connection_button))
test_connection_button.grid(row=3, columnspan=2, pady=20)

# Create StringVar instances to hold folder paths
input_folder_var = tk.StringVar()
output_folder_var = tk.StringVar()

# Folder selection widgets
input_label = tk.Label(root, text="Select Input Folder:")
input_label.grid(row=4, column=0, padx=10, pady=5)
input_entry = tk.Entry(root, textvariable=input_folder_var, width=50)
input_entry.grid(row=4, column=1, padx=10, pady=5)
input_button = tk.Button(root, text="Browse", command=select_input_folder)
input_button.grid(row=4, column=2, padx=10, pady=5)

output_label = tk.Label(root, text="Select Output Folder:")
output_label.grid(row=5, column=0, padx=10, pady=5)
output_entry = tk.Entry(root, textvariable=output_folder_var, width=50)
output_entry.grid(row=5, column=1, padx=10, pady=5)
output_button = tk.Button(root, text="Browse", command=select_output_folder)
output_button.grid(row=5, column=2, padx=10, pady=5)

#show_button = tk.Button(root, text="Show Selected Folders", command=show_selected_folders)
#show_button.grid(row=6, columnspan=3, pady=10)

task_button = tk.Button(root, text="Perform Statistic Analysis", command=start_long_running_task)
task_button.grid(row=7, columnspan=3, pady=20)

# Start the Tkinter event loop
root.mainloop()
